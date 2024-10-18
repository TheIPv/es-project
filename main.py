import os
import datetime
import tkinter as tk
import threading
from tkinter import filedialog, messagebox
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Отключает отображение графиков на экране


# Глобальные переменные для хранения загруженных данных
df = None
OSV = None

# Функция для создания уникальной папки на рабочем столе
def create_output_folder():
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    folder_path = os.path.join(desktop, f'Анализ_{timestamp}')
    os.makedirs(folder_path)
    return folder_path

# Функция для обновления лога
def update_log(message):
    log_label.config(text=message)
    window.update_idletasks()

# Функция для загрузки файла
def load_file(file_type):
    update_log(f"Файл {file_type} загружается...")
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        messagebox.showwarning("Ошибка", f"Файл {file_type} не выбран!")
        return None

    if file_type == "ЖО":
        global df
        df = pd.read_excel(file_path, converters={'Счет Кт': str, 'Счет Дт': str})
    else:  # file_type == "ОСВ"
        global OSV
        OSV = pd.read_excel(file_path).round(decimals=2).set_index('Счет')

    update_log(f"Файл {file_type} загружен")
    return file_path

# Функция для загрузки первого файла
def load_first_file():
    load_file("ЖО")
    load_osv_button.config(state=tk.NORMAL)

# Функция для загрузки второго файла
def load_second_file():
    global df, OSV
    load_file("ОСВ")
    if df is not None and OSV is not None:
        update_log("Файлы загружены. Готово к анализу.")
        start_analysis_button.pack(pady=10)

def analyze_data_thread():
    analyze_data()
def start_analysis():
    threading.Thread(target=analyze_data_thread).start()

# Функция для анализа данных
def analyze_data():
    global df, OSV
    import numpy as np
    import pandas as pd
    import matplotlib.pyplot as plt
    import tkinter as tk
    import os
    from tkinter import filedialog, messagebox
    update_log("Тест на целостность данных...")
    window.update_idletasks()  # Обновляем интерфейс

    # Тест целостности данных
    names = '000, 001, 002, 003, 003.01, 003.02, 004, 004.1, 004.02, 004.К, 005, 006, 007, 008, 008.1, 008.21, 009, 009.01, 009.21, 010, 011, 012, 012.01, 012.02, ГТД, КВ, МЦ, МЦ.02, МЦ.03, МЦ.04, НЕ, НЕ.01, НЕ.01.1, НЕ.01.9, НЕ.02, НЕ.02.1, НЕ.02.9, НЕ.03, НЕ.04, ОТ, ОТ.01, ОТ.02, ОТ.03, РВ, РВ.1, РВ.2, РВ.3, РВ.4, УСН, УСН.01, УСН.01, УСН.02, УСН.03, УСН.04, УСН.21, УСН.22, УСН.23, УСН.24, Я75, Я81.01, Я81, Я80, Я80.02, Я80.01, Я80.09, Я75.01, Я81.09, Я81.02, Я75.02, Я69.06.5, Я01.К, Я96, Я96.01'
    namesofaccs = [accs.strip() for accs in names.split(',')]
    result = df[df['СчетДт'].isnull() | df['СчетКт'].isnull()]
    result = result.loc[~result['СчетКт'].isin(namesofaccs) & ~result['СчетДт'].isin(namesofaccs)]

    if result.shape[0] == 0:
        print("Проверка целостности данных успешно завершена")
    else:
        print("Проверка целостности данных выполнена с ошибками")
    update_log("Тест математической правильности...")
    # Тест математической правильности
    resultosv = OSV.loc[(OSV['Сальдо начальное Дт'].fillna(0) - OSV['Сальдо начальное Кт'].fillna(0) +
                         OSV['Обороты Дт'].fillna(0) - OSV['Обороты Кт'].fillna(0) -
                         OSV['Сальдо конечное Дт'].fillna(0) + OSV['Сальдо конечное Кт'].fillna(0)).abs() > 0.9]

    if resultosv.shape[0] == 0:
        print("Проверка математической правильности успешно завершена")
    else:
        print("Проверка математической правильности выполнена с ошибками")
    update_log("Тест полноты выгрузки...")
    # Тест полноты выгрузки
    journal_sum = df[['СчетДт', 'СчетКт', 'Сумма']].fillna(0)
    journal_sum_dt = journal_sum.groupby(['СчетДт'], as_index=False)['Сумма'].sum()
    journal_sum_kt = journal_sum.groupby(['СчетКт'], as_index=False)['Сумма'].sum()
    journal_sum_kt = journal_sum_kt.rename(columns={'Сумма': 'Обороты Кт по выгрузке'})
    journal_sum_dt = journal_sum_dt.rename(columns={'Сумма': 'Обороты Дт по выгрузке'}).set_index('СчетДт')
    journal_sum_kt = journal_sum_kt.set_index('СчетКт')
    osvjournal = pd.concat([OSV, journal_sum_kt.reindex(OSV.index)], axis=1)
    osvjournal = pd.concat([osvjournal, journal_sum_dt.reindex(OSV.index)], axis=1)
    osvjournal['Обороты Дт разница'] = osvjournal['Обороты Дт'].round(2) - osvjournal['Обороты Дт по выгрузке'].round(2)
    osvjournal['Обороты Кт разница'] = osvjournal['Обороты Кт'].round(2) - osvjournal['Обороты Кт по выгрузке'].round(2)
    osvresult = osvjournal.loc[(osvjournal['Обороты Дт разница'] != 0) | (osvjournal['Обороты Кт разница'] != 0)]

    if len(osvresult) != 0:
        print("Проверка полноты выгрузки выполнена с ошибками")
    else:
        print("Проверка полноты выгрузки успешно завершена")

    """# **Проверка законом Бенфрода**

                ## **Вспомогательный код**
                """
    update_log("Проведение вспомогательных вычислений...")
    import ru_benford as bf
    benf = bf.Benford(df['Сумма'], decimals=2, confidence=99)

    benf.summation()

    # Добавляем признак: является ли сумма сторнированной (1) или нет (0)
    df["Сторно"] = df["Сумма"].apply(lambda x: 0 if x > 0 else 1)

    # Убираем строки с пустыми значениями поля Сумма
    df = df[df['Сумма'].notnull()]

    # Для отрицательных сумм меняем знак
    df.loc[df["Сумма"] < 0, "Сумма"] = -df["Сумма"]

    # Оставляем значения, больше 10, чтобы на них можно было провести все тесты
    df = df[df['Сумма'] > 10]

    # Меняем индексы
    df = df.reset_index(drop=True)
    """## **2.1 Базовые тесты**

    ### **2.1.1 Тест первой цифры**
    """

    benf.F1D.report()

    """### **2.1.2 Тест второй цифры**"""

    benf.SD.report()

    """### **2.1.3 Тест первых двух цифр**"""

    benf.F2D.report()

    """## **2.2 Продвинутые тесты**

    ### **2.2.1 Тест суммирования**
    """

    sm = bf.summation(df['Сумма'], decimals=2)

    """### **2.2.2 Тест второго порядка**"""

    benf.sec_order()
    benf.F2D_sec.report()

    """### **2.2.3 Тест мантисс**"""

    mant = bf.mantissas(df['Сумма'])

    """## **2.3 Связанные тесты**

    ### **2.3.1 Тест дублирования сумм**
    """

    # Считаем частоту появления сумм

    temp = df.groupby('Сумма')['Сумма'].count() / len(df)
    sum_ = list(temp.index)
    frequency = list(temp)
    df_temp = pd.DataFrame({"Сумма": sum_, "Частота суммы": frequency})

    df = df.merge(df_temp, on='Сумма')
    test_overwrite = df
    most_frequent_values = df['Сумма'].value_counts().head(10)

    sorted_counts = most_frequent_values.sort_values(ascending=False)

    bar_width = 0.5  # Увеличиваем ширину столбцов
    bar_positions = [i + bar_width / 2 for i in range(len(sorted_counts))]  # Смещаем позиции столбцов

    plt.figure(figsize=(10, 6))  # Увеличиваем размер графика
    plt.bar(bar_positions, sorted_counts.values, width=bar_width, align='center',
            color=(0 / 255, 121 / 255, 140 / 255))  # Устанавливаем цвет в RGB
    plt.xlabel('Значения')
    plt.ylabel('Частота')
    plt.title('Тест дублирования сумм')
    plt.xticks(bar_positions, sorted_counts.index, rotation=45)  # Вращаем подписи для лучшей читаемости
    plt.tight_layout()  # Оптимально размещаем элементы на графике
    #plt.show()

    """### **2.3.2 Тест двух последних цифр**"""

    benf.L2D.report()

    """### **2.3.3 Оценка коэффициента искажения**"""

    result = df.loc[df['Сумма'] >= 10]
    result.loc[result['Сумма'] != 0, ['Сумма']] = (
            10 * result['Сумма'] / (10 ** (np.log10(result['Сумма']).fillna(0)).astype(int)))

    Avg_Found = result['Сумма'].sum() / result['Сумма'].count()
    Avg_Expected = 90 / (result['Сумма'].count() * (10 ** (1 / result['Сумма'].count()) - 1))
    Distortion = (Avg_Found - Avg_Expected) / Avg_Expected * 100
    Std = result['Сумма'].std()
    Z_stat = Distortion / Std

    print(f'Среднее Факт: {round(Avg_Found, 2)}')
    print(f'Среднее Теор: {round(Avg_Expected, 2)}')
    print(f'Коэффициент искажения: {round(Distortion, 2)}%')
    print(f'Z-статистика: {round(Z_stat, 2)}')
    print('Критическое значение: 2,57')

    if (Distortion < 0):
        print('Вывод: Значения в массиве занижены')
    else:
        print('Вывод: Значения в массиве завышены')
    if (abs(Z_stat) > 2, 57):
        print('Коэффициент искажения является существенным')
    else:
        print('Коэффициент искажения является несущественным')

    # Оценка риска

    """# ------ [ ВЫГРУЗКА РЕЗУЛЬТАТОВ В EXCEL ] ------"""
    from openpyxl import load_workbook, Workbook
    from openpyxl.drawing.image import Image
    # 1. Закон Бенфорда

    update_log("Выгрузка законов Бенфорда...")
    benf.summation()

    initial_size = len(benf.chosen)
    tested_sample_size = len(benf.base)
    data = (
        f"'F1D': {benf._discarded['F1D']}\n; "
        f"'F2D': {benf._discarded['F2D']}\n; "
        f"'F3D': {benf._discarded['F3D']}\n; "
        f"'SD': {benf._discarded['SD']}\n; "
        f"'L2D': {benf._discarded['L2D']};"
    )
    # Путь к папке для сохранения отчетов
    output_folder = create_output_folder()
    # Путь для сохранения Excel-файла
    output_path = os.path.join(output_folder, 'Отчёт.xlsx')

    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()

    if 'Результаты тестов' in workbook.sheetnames:
        worksheet = workbook['Результаты тестов']
    else:
        # Если такого листа нет, создаем новый
        worksheet = workbook.create_sheet('Результаты тестов')

    # Записываем данные в существующий или новый лист
    worksheet['A1'] = 'Исходный размер выборки'
    worksheet['B1'] = len(benf.chosen)

    worksheet['A2'] = 'Количество значений, на которых проведен тест'
    worksheet['B2'] = len(benf.base)

    data = (
        f"'F1D': {benf._discarded['F1D']}\n; "
        f"'F2D': {benf._discarded['F2D']}\n; "
        f"'F3D': {benf._discarded['F3D']}\n; "
        f"'SD': {benf._discarded['SD']}\n; "
        f"'L2D': {benf._discarded['L2D']};"
    )
    worksheet['A3'] = 'Исключенные значения для каждого теста'
    worksheet['B3'] = data

    workbook.save(output_path)

    print(f'Отчет сохранен в файле: {output_path}')
    update_log("Выгрузка тестов первой, второй и 1 и 2 цифры...")

    """## **3. Тест первой, второй и 1и2 цифры**"""

    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    import io
    import sys

    # Путь к файлу Excel
    file_name = os.path.join(output_folder, 'Отчёт.xlsx')

    # Запись данных в Excel
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a' if os.path.exists(output_path) else 'w') as writer:
        for name, test, sheet_title in [
            ('F1D', benf.F1D, 'Тест первой цифры'),
            ('SD', benf.SD, 'Тест второй цифры'),
            ('F2D', benf.F2D, 'Тест 1 и 2 цифры')
        ]:
            output_buffer = io.StringIO()

            sys.stdout = output_buffer

            # Генерация отчета теста без отображения графика
            report_data = test.report(show_plot=False)

            sys.stdout = sys.__stdout__

            captured_output = output_buffer.getvalue()
            output_buffer.close()

            # Преобразование вывода в DataFrame
            output_df = pd.DataFrame([line] for line in captured_output.splitlines())
            output_df.to_excel(writer, sheet_name=sheet_title, header=False, index=False, startrow=0)

            if report_data is not None:
                if isinstance(report_data, pd.DataFrame):
                    report_data.to_excel(writer, sheet_name=sheet_title, index=True,
                                         startrow=len(output_df) + 2)

    # Открытие существующего файла Excel для добавления изображений
    workbook = load_workbook(file_name)

    # Добавление графиков к соответствующим листам
    for name, sheet_title in [('F1D', 'Тест первой цифры'), ('SD', 'Тест второй цифры'), ('F2D', 'Тест 1 и 2 цифры')]:
        plot_file = f'{name}_plot.png'
        test = getattr(benf, name)
        test.report(show_plot=True, save_plot=plot_file)  # Сохранение графика в файл

        worksheet = workbook[sheet_title]
        img = Image(plot_file)

        # Настройка размера изображения (уменьшение до 50% от оригинального)
        img.width = img.width // 2
        img.height = img.height // 2

        # Определение позиции изображения (правее таблицы)
        max_column = worksheet.max_column  # Находим последнюю колонку с данными
        image_column = max_column + 2  # Смещаемся на 2 колонки вправо

        # Добавляем изображение в нужную позицию
        worksheet.add_image(img, f'{chr(65 + image_column)}5')  # Столбец A + image_column

    # Сохранение изменений в файле Excel
    workbook.save(file_name)

    print(f"Отчеты, таблицы и графики успешно сохранены в файл {file_name}")

    """## **4. Тест суммирования**"""

    """## **5. Тест второго порядка**"""
    import os
    import io
    import sys
    import re
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

    # Путь к файлу
    file_name = os.path.join(output_folder, 'Отчёт.xlsx')

    # Выгрузка теста второго порядка
    update_log("Выгрузка теста второго порядка...")

    # Перехватываем вывод команды benf.F2D_sec.report()
    benf.sec_order()

    # Создаем строковый буфер для перехвата вывода
    report_output = io.StringIO()
    sys.stdout = report_output

    # Выполняем команду, которая выводит результаты
    image_path = 'plot.png'
    benf.F2D_sec.report(save_plot=image_path)

    # Возвращаем стандартный вывод обратно
    sys.stdout = sys.__stdout__

    # Получаем содержимое вывода как строку
    report_text = report_output.getvalue()

    # Парсим данные с помощью регулярных выражений для создания таблицы
    pattern = re.compile(r'(\d{1,2})\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)')
    parsed_data = pattern.findall(report_text)

    # Создаем списки для DataFrame с результатами
    first_two_digits = []
    expected = []
    found = []
    z_scores = []

    for entry in parsed_data:
        first_two_digits.append(entry[0])
        expected.append(float(entry[1]))
        found.append(float(entry[2]))
        z_scores.append(float(entry[3]))

    # Создаем DataFrame с четкими заголовками
    df_results = pd.DataFrame({
        'Первые цифры': first_two_digits,
        'Теор': expected,
        'Факт': found,
        'Z_статистика': z_scores
    })

    # Удаляем строки таблицы из полного текста отчета и дублирующиеся заголовки
    report_lines = [line.strip() for line in report_text.splitlines() if not pattern.match(line)]
    report_lines_cleaned = []
    skip_next = False

    for line in report_lines:
        if "Теор" in line and "Факт" in line and "Z_статистика" in line:
            if not skip_next:
                skip_next = True
                continue
        if skip_next:
            skip_next = False
            continue
        report_lines_cleaned.append(line)

    # Открытие существующего файла Excel
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
    else:
        workbook = Workbook()

    # Создаем новый лист для теста второго порядка
    sheet_name = 'Тест второго порядка'
    if sheet_name not in workbook.sheetnames:
        worksheet_results = workbook.create_sheet(sheet_name)
    else:
        worksheet_results = workbook[sheet_name]

    # Записываем очищенный полный отчет, разбивая строки по пробелам
    for row_num, line in enumerate(report_lines_cleaned):
        columns = line.split()
        for col_num, cell_value in enumerate(columns):
            worksheet_results.cell(row=row_num + 1, column=col_num + 1, value=cell_value)

    # Определяем начальную строку для вставки DataFrame
    start_row = len(report_lines_cleaned) + 2  # Добавляем небольшой отступ

    # Записываем DataFrame вниз под текстовым отчетом
    for col_num, header in enumerate(df_results.columns):
        worksheet_results.cell(row=start_row, column=col_num + 1, value=header)  # Заголовки

    for row_num, (digit, exp, found_val, z) in enumerate(
            zip(df_results['Первые цифры'], df_results['Теор'], df_results['Факт'],
                df_results['Z_статистика']),
            start=start_row + 1):
        worksheet_results.cell(row=row_num, column=1, value=digit)  # Первые цифры
        worksheet_results.cell(row=row_num, column=2, value=exp)  # Теор
        worksheet_results.cell(row=row_num, column=3, value=found_val)  # Факт
        worksheet_results.cell(row=row_num, column=4, value=z)  # Z-статистика

    # Вставляем график в Excel на новый лист
    img = Image(image_path)
    img.width = img.width // 2  # Уменьшаем размер изображения
    img.height = img.height // 2
    worksheet_results.add_image(img, f'K1')  # Вставляем график в ячейку K1

    # Сохранение изменений в файл Excel
    workbook.save(file_name)

    print(f"Отчет сохранен в файл {file_name}")

    # Удаляем временный файл изображения
    import os
    os.remove(image_path)

    """## **6. Тест мантисс**"""

    # Путь к файлу Excel
    file_name = os.path.join(output_folder, 'Отчёт.xlsx')

    # Выгрузка теста мантисс
    update_log("Выгрузка теста мантисс...")

    # Перехватываем вывод команды
    report_output = io.StringIO()
    sys.stdout = report_output

    # Закрываем все открытые фигуры
    plt.close('all')

    # Генерируем графики и текст
    mant = bf.mantissas(df['Сумма'])

    # Возвращаем стандартный вывод обратно
    sys.stdout = sys.__stdout__

    # Получаем содержимое вывода как строку
    report_text = report_output.getvalue()
    report_lines = [line.strip() for line in report_text.splitlines()]

    # Сохраняем все открытые графики
    image_path_1 = 'plot_1.png'
    image_path_2 = 'plot_2.png'
    figures = [plt.figure(i) for i in plt.get_fignums()]

    # Сохраняем графики
    if len(figures) >= 2:
        figures[0].savefig(image_path_1)
        figures[1].savefig(image_path_2)

    # Закрываем все открытые фигуры
    plt.close('all')

    # Открываем или создаем файл Excel
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
    else:
        workbook = Workbook()

    # Добавляем новый лист для теста Мантисс
    sheet_name = 'Тест Мантисс'
    if sheet_name not in workbook.sheetnames:
        worksheet_results = workbook.create_sheet(sheet_name)
    else:
        worksheet_results = workbook[sheet_name]

    # Записываем очищенный полный отчет, разбивая строки по пробелам
    for row_num, line in enumerate(report_lines):
        columns = line.split()
        for col_num, cell_value in enumerate(columns):
            worksheet_results.cell(row=row_num + 1, column=col_num + 1, value=cell_value)

    # Вставляем графики в Excel
    if os.path.exists(image_path_1):
        img_1 = Image(image_path_1)
        img_1.width, img_1.height = img_1.width // 2, img_1.height // 2
        worksheet_results.add_image(img_1, 'A12')  # Вставляем первый график

    if os.path.exists(image_path_2):
        img_2 = Image(image_path_2)
        img_2.width, img_2.height = img_2.width // 2, img_2.height // 2
        worksheet_results.add_image(img_2, 'K1')  # Вставляем второй график

    # Сохраняем изменения в файл Excel
    workbook.save(file_name)

    # Удаляем временные файлы изображений
    if os.path.exists(image_path_1):
        os.remove(image_path_1)
    if os.path.exists(image_path_2):
        os.remove(image_path_2)

    print(f"Отчет и графики успешно сохранены в файл {file_name}")

    """## **7. Связанные тесты**"""

    update_log("Выгрузка связанных тестов...")
    # Путь к файлу Excel
    output_path = os.path.join(output_folder, 'Отчёт.xlsx')

    # Проверка существования файла
    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()

    "# 7.1 Тест дублирования сумм"

    update_log("Выгрузка теста дублирования сумм...")
    temp = df.groupby('Сумма')['Сумма'].count() / len(df)
    sum_ = list(temp.index)
    frequency = list(temp)
    df_temp = pd.DataFrame({"Сумма": sum_, "Частота суммы": frequency})
    df = df.merge(df_temp, on='Сумма')
    most_frequent_values = df['Сумма'].value_counts().head(10)
    sorted_counts = most_frequent_values.sort_values(ascending=False)

    # Построение графика
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar([i + 0.5 for i in range(len(sorted_counts))], sorted_counts.values, width=0.5, align='center',
           color=(0 / 255, 121 / 255, 140 / 255))
    ax.set_xlabel('Значения')
    ax.set_ylabel('Частота')
    ax.set_title('Тест дублирования сумм')
    ax.set_xticks([i + 0.5 for i in range(len(sorted_counts))])
    ax.set_xticklabels(sorted_counts.index, rotation=45)

    plt.tight_layout()
    image_path = 'plot_duplication.png'
    fig.savefig(image_path)
    plt.close(fig)

    # Добавление листа для теста дублирования сумм
    sheet_name = 'Тест дублирования сумм'
    if sheet_name not in workbook.sheetnames:
        worksheet_results = workbook.create_sheet(sheet_name)
    else:
        worksheet_results = workbook[sheet_name]

    # Вставка графика
    img_duplication = Image(image_path)
    img_duplication.width, img_duplication.height = img_duplication.width // 2, img_duplication.height // 2
    worksheet_results.add_image(img_duplication, 'A1')

    # Сохранение изменений
    workbook.save(output_path)

    # Удаление временного файла изображения
    os.remove(image_path)

    "# 7.2 Тест двух последних цифр"
    update_log("Выгрузка теста двух последних цифр...")

    # Перехватываем вывод команды benf.F2D_sec.report()
    benf.sec_order()

    # Создаем строковый буфер для перехвата вывода
    report_output = io.StringIO()
    sys.stdout = report_output

    # Выполняем команду, которая выводит результаты
    image_path = 'plot.png'
    benf.L2D.report(save_plot=image_path)

    # Возвращаем стандартный вывод обратно
    sys.stdout = sys.__stdout__

    # Получаем содержимое вывода как строку
    report_text = report_output.getvalue()

    # Парсим данные с помощью регулярных выражений для создания таблицы
    pattern = re.compile(r'(\d{1,2})\s+([0-9.]+)\s+([0-9.]+)\s+([0-9.]+)')
    parsed_data = pattern.findall(report_text)

    # Создаем списки для DataFrame с результатами
    first_two_digits = []
    expected = []
    found = []
    z_scores = []

    for entry in parsed_data:
        first_two_digits.append(entry[0])
        expected.append(float(entry[1]))
        found.append(float(entry[2]))
        z_scores.append(float(entry[3]))

    # Создаем DataFrame с четкими заголовками
    df_results = pd.DataFrame({
        'Первые цифры': first_two_digits,
        'Теор': expected,
        'Факт': found,
        'Z_статистика': z_scores
    })

    # Удаляем строки таблицы из полного текста отчета и дублирующиеся заголовки
    report_lines = [line.strip() for line in report_text.splitlines() if not pattern.match(line)]
    report_lines_cleaned = []
    skip_next = False

    for line in report_lines:
        if "Теор" in line and "Факт" in line and "Z_статистика" in line:
            if not skip_next:
                skip_next = True
                continue
        if skip_next:
            skip_next = False
            continue
        report_lines_cleaned.append(line)

    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()

    # Записываем данные в Excel
    worksheet_results = workbook.create_sheet('Тест двух последних цифр')

    # Записываем очищенный полный отчет, разбивая строки по пробелам
    for row_num, line in enumerate(report_lines_cleaned):
        columns = line.split()
        for col_num, cell_value in enumerate(columns):
            worksheet_results.cell(row=row_num + 1, column=col_num + 1, value=cell_value)

    # Определяем начальную строку для вставки DataFrame
    start_row = len(report_lines_cleaned) + 2  # Добавляем небольшой отступ

    # Записываем DataFrame вниз под текстовым отчетом
    for col_num, header in enumerate(df_results.columns):
        worksheet_results.cell(row=start_row, column=col_num + 1, value=header)  # Записываем заголовки

    for row_num, (digit, exp, found_val, z) in enumerate(
            zip(df_results['Первые цифры'], df_results['Теор'], df_results['Факт'], df_results['Z_статистика']),
            start=start_row + 1):
        worksheet_results.cell(row=row_num, column=1, value=digit)  # Первые цифры
        worksheet_results.cell(row=row_num, column=2, value=exp)  # Теор
        worksheet_results.cell(row=row_num, column=3, value=found_val)  # Факт
        worksheet_results.cell(row=row_num, column=4, value=z)  # Z-статистика

    from openpyxl.drawing.image import Image
    img = Image(image_path)
    worksheet_results.add_image(img, 'K1')

    workbook.save(output_path)

    # Удаляем временный файл изображения
    import os
    os.remove(image_path)

    "# 7.3 Оценка коэффициента искажения"

    import pandas as pd
    import numpy as np

    # Предположим, что df уже определен. Ваш расчет значений:
    result = df.loc[df['Сумма'] >= 10]
    result.loc[result['Сумма'] != 0, ['Сумма']] = (
            10 * result['Сумма'] / (10 ** (np.log10(result['Сумма']).fillna(0)).astype(int)))

    Avg_Found = result['Сумма'].sum() / result['Сумма'].count()
    Avg_Expected = 90 / (result['Сумма'].count() * (10 ** (1 / result['Сумма'].count()) - 1))
    Distortion = (Avg_Found - Avg_Expected) / Avg_Expected * 100
    Std = result['Сумма'].std()
    Z_stat = Distortion / Std

    # Подготовка данных для записи в Excel
    data = {
        'Показатель': [
            'Среднее Факт', 'Среднее Теор', 'Коэффициент искажения', 'Z-статистика', 'Критическое значение',
            'Вывод по значениям', 'Значимость искажения'
        ],
        'Значение': [
            round(Avg_Found, 2), round(Avg_Expected, 2), f"{round(Distortion, 2)}%", round(Z_stat, 2), 2.57,
            'Значения в массиве занижены' if Distortion < 0 else 'Значения в массиве завышены',
            'Коэффициент искажения является существенным' if abs(
                Z_stat) > 2.57 else 'Коэффициент искажения является несущественным'
        ]
    }

    df_output = pd.DataFrame(data)

    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
    else:
        workbook = Workbook()

    # Добавляем новый лист для анализа искажения
    worksheet = workbook.create_sheet('Анализ искажения')

    # Запись данных DataFrame в лист
    for row_num, (indicator, value) in enumerate(zip(df_output['Показатель'], df_output['Значение']), start=2):
        worksheet.cell(row=row_num, column=1, value=indicator)  # Показатель
        worksheet.cell(row=row_num, column=2, value=value)  # Значение

    # Установка ширины столбцов по максимальной длине значений
    for col_num, col_name in enumerate(df_output.columns, start=1):
        max_len = max(df_output[col_name].astype(str).map(len).max(), len(col_name)) + 2
        worksheet.column_dimensions[chr(64 + col_num)].width = max_len

    # Добавление заголовка
    worksheet.cell(row=1, column=1, value='Анализ искажения данных')
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    workbook.save(output_path)

    """Кластеризация. Этап 1"""
    update_log("Подготовка к кластеризации...")
    df = test_overwrite
    print(df)
    import math
    from datetime import datetime
    # Вероятность появления цифры в первом разряде

    def p_first(d1):
        return math.log10(1 + 1 / d1)

    # Вероятность появления цифры во втором разряде

    def p_second(d2):
        s = 0
        for k in range(1, 10):
            s += math.log10(1 + 1 / (10 * k + d2))
        return s

    # Список, содержащий вероятности появления цифр 1-9 в первом #разряде

    first_teor = [p_first(d1) for d1 in range(1, 10)]
    df_duplicates = df.groupby(["СчетДт", "СчетКт", "Сумма"], as_index=False).count().sort_values(
        by="Организация",
        ascending=False)

    df.loc[:, 'first'] = df['Сумма'].apply(lambda x: int(str(x)[0]))

    # Список, содержащий частоты появления первых цифр в выборке

    first_real = df.groupby('first')['Сумма'].count() / len(df)

    # Расчет среднего абсолютного отклонения

    def MAD(AP, EP, k):
        s = 0
        for i in range(0, k - 1):
            s += abs(AP[i] - EP[i])
        return s / k

    mad = MAD(list(first_real), first_teor, 9)

    # Z-статистика

    def z_stat(AP, EP, N):
        chisl = abs(AP - EP)
        znam = ((EP * (1 - EP)) / N) ** 0.5
        if 1 / (2 * N) < chisl:
            chisl -= 1 / (2 * N)
        return chisl / znam

    # Z-Тест 1 цифры

    z_stats = []
    for i in range(9):
        z_stats.append(z_stat(list(first_real)[i], first_teor[i], len(df)))

    # Расчет хи-квадрат

    def chi2(AC, EC, N):
        k = len(AC)
        chi = 0
        for i in range(k):
            chi += (AC[i] * N - EC[i] * N) ** 2 / EC[i] * N
        return chi

    chi_stat = chi2(list(first_real), first_teor, len(df))

    # Добавление в исходный датафрейм столбца со значениями z-#статистик

    df_first_stats = pd.DataFrame({"first": list(range(1, 10)), "z-stat first": z_stats})
    df = df.merge(df_first_stats, on='first')

    # Z-Тест 2 цифры

    df.loc[:, 'second'] = df['Сумма'].apply(lambda x: int(str(x)[1]))

    second_real = df.groupby('second')['Сумма'].count() / len(df)

    second_teor = [p_second(d1) for d1 in range(0, 10)]

    z_stat_sec = []
    for i in range(10):
        z_stat_sec.append(z_stat(list(second_real)[i], second_teor[i], len(df)))

    df_second_stats = pd.DataFrame({"second": list(range(0, 10)), "z-stat second": z_stat_sec})
    df = df.merge(df_second_stats, on='second')

    # Тест первых двух цифр

    df.loc[:, 'first_two'] = df['Сумма'].apply(lambda x: int(str(x)[:2]))

    two_teor = [p_first(d1) for d1 in range(10, 100)]
    two_real = df.groupby('first_two')['Сумма'].count() / len(df)

    z_stat_two = []
    for i in range(90):
        z_stat_two.append(z_stat(list(two_real)[i], two_teor[i], len(df)))

    df_first_two_stats = pd.DataFrame({"first_two": list(range(10, 100)), "z-stat first_two": z_stat_two})
    df = df.merge(df_first_two_stats, on='first_two')

    # Тест суммирования

    two_real = df.groupby('first_two')['Сумма'].sum() / df['Сумма'].sum()

    df_abs_delta = pd.DataFrame({"first_two": list(range(10, 100)), "sum_frequency": list(two_real)})
    df = df.merge(df_abs_delta, on='first_two')

    # Тест второго порядка

    df_cur = df.sort_values(by='Сумма')
    df_cur.loc[:, 'two'] = df_cur['Сумма'].diff() * 10
    df_cur.dropna(subset=['Сумма'], inplace=True)
    df_cur = df_cur[df_cur['two'] > 10]
    df_cur.loc[:, 'two'] = df_cur['two'].apply(lambda x: int(str(x)[:2]))
    df_cur.shape

    df_z_stat_second_diff = pd.DataFrame({"two": list(range(10, 100)), "z_stat_second_diff": z_stat_two})

    df_cur.head()

    df_cur = df_cur.merge(df_z_stat_second_diff, on="two")

    ind = df_cur.index
    df.loc[ind, "z_stat_second_diff"] = df_cur["z_stat_second_diff"]

    # Z-Тест последних двух цифр

    df_cur = df

    df_cur.loc[:, 'last_two'] = df_cur['Сумма'].apply(lambda x: int(str(int(round((x * 100), 0)))[-2:]))

    two_real = df_cur.groupby('last_two')['Сумма'].count() / len(df_cur)

    two_teor = [0.01 for i in range(100)]

    z_stats = []
    for i in range(100):
        z_stats.append(z_stat(list(two_real)[i], two_teor[i], len(df_cur)))

    mad = MAD(list(two_real), two_teor, 100)

    df_last_two = pd.DataFrame({"last_two": list(range(0, 100)), "z_stat_last_two": z_stats})
    df_cur = df_cur.merge(df_last_two, on='last_two')

    def a_socr(a):
        return 10 * a / (10 ** int(math.log(a, 10)))

    df_cur['two'] = df_cur['Сумма'].apply(lambda x: a_socr(x))

    # Добавляется столбец с частотой счета Дт (синтетический счет)

    df_cur["СинтСчетДт"] = df_cur["СчетДт"].apply(lambda x: str(x)[:2])
    df_dt = (df_cur.groupby("СинтСчетДт").count() / len(df_cur))
    df_dt = df_dt.rename(columns={"Организация": "Частота счета Дт"})
    df_dt = df_dt["Частота счета Дт"]
    df_cur = df_cur.merge(df_dt, on='СинтСчетДт')

    # Добавляется столбец с частотой счета Кт (синтетический счет)

    df_cur["СинтСчетКт"] = df_cur["СчетКт"].apply(lambda x: str(x)[:2])
    df_kt = (df_cur.groupby("СинтСчетКт").count() / len(df_cur))
    df_kt = df_kt.rename(columns={"Организация": "Частота счета Кт"})
    df_kt = df_kt["Частота счета Кт"]
    df_cur = df_cur.merge(df_kt, on='СинтСчетКт')

    # Добавляется столбец с частотой проводки (по синтетическим счетам)

    df_cur["Проводка"] = df_cur["СинтСчетДт"] + "-" + df_cur["СинтСчетКт"]
    df_pr = (df_cur.groupby("Проводка").count() / len(df_cur))
    df_pr = df_pr.rename(columns={"Организация": "Частота проводки"})
    df_pr = df_pr["Частота проводки"]
    df_cur = df_cur.merge(df_pr, on="Проводка")

    # Добавляется столбец с частотой автора операции

    df_au = df_cur.groupby("АвторОперации").count() / len(df_cur)
    df_au = df_au.rename(columns={"Организация": "Частота автора операции"})
    df_au = df_au["Частота автора операции"]
    df_cur = df_cur.merge(df_au, on='АвторОперации')

    # Ручная проводка (1) или нет (0)

    df_cur["Ручная проводка"] = df_cur["РучнаяКорректировка"].apply(lambda x: 1 if x == "Да" else 0)

    # выходные дни (1), другие дни (0)

    df_cur["Data"] = df_cur["Период"].apply(lambda x: str(x)[:10])
    df_cur["Data"] = df_cur["Data"].apply(lambda x: datetime.strptime(x, "%Y-%m-%d"))
    df_cur["Data"] = df_cur["Data"].apply(lambda x: datetime.weekday(x))
    df_cur["Data"] = df_cur["Data"].apply(lambda x: 1 if (x == 5 or x == 6) else 0)
    df_cur = df_cur.rename(columns={"Data": "Выходные или рабочие"})

    # Считает количество дублирующихся проводок

    df_duplicates_ = df_duplicates.iloc[:, :4]
    df_duplicates_.rename({"Организация": "Количество дублей"}, axis=1, inplace=True)
    df_cur = df_cur.merge(df_duplicates_, on=["СчетДт", "СчетКт", "Сумма"])

    temp = df_cur.loc[:,
           ["z-stat first", "z-stat second", "z-stat first_two", "sum_frequency", "z_stat_second_diff",
            "z_stat_last_two",
            "Частота суммы", "Частота счета Дт", "Частота счета Кт", "Частота проводки",
            "Частота автора операции",
            "Ручная проводка", "Выходные или рабочие", "Сторно", "Количество дублей"]]

    temp.loc[temp["z_stat_second_diff"].isna(), "z_stat_second_diff"] = -1
    # Считает количество дублирующихся проводок
    df_duplicates_ = df_duplicates.iloc[:, :4]
    df_duplicates_.rename({"Организация": "Количество дублей"}, axis=1, inplace=True)
    df_cur = df_cur.merge(df_duplicates_, on=["СчетДт", "СчетКт", "Сумма"])

    temp.head()
    from sklearn import preprocessing
    scaled = preprocessing.StandardScaler().fit_transform(temp.iloc[:, :6].values)
    scaled_df = pd.DataFrame(scaled, index=temp.iloc[:, :6].index, columns=temp.iloc[:, :6].columns)
    scaled_df.head()
    plt.close('all')  # Закрывает все открытые графики

    import os
    import matplotlib.pyplot as plt
    from tqdm import tqdm
    from sklearn.cluster import KMeans
    import sklearn.metrics as metrics
    update_log("Построение силуэта...")
    # Настройки для графиков
    plt.rcParams["figure.figsize"] = (8, 4)

    # Список возможных значений кластеров
    x = [i for i in range(2, 6)]
    m = []

    sample_size = int(0.2 * scaled_df.shape[0])
    print(f'sample size = {sample_size}')

    # Подсчёт silhouette_score для каждого количества кластеров
    for i in tqdm(x):
        labels = KMeans(n_clusters=i, random_state=1000).fit(scaled_df).labels_
        m.append(metrics.silhouette_score(scaled_df, labels, sample_size=sample_size))

    # Найдём количество кластеров с максимальным значением silhouette_score
    best_n_clusters = x[m.index(max(m))]

    print(f'Оптимальное количество кластеров: {best_n_clusters}')

    # Строим график зависимости silhouette_score от количества кластеров
    plt.plot(x, m, 'r-')
    plt.xticks(ticks=x, labels=[int(i) for i in x])
    plt.xlabel('Количество кластеров')
    plt.ylabel('Значение метрики')
    plt.title('Зависимость значения метрики от количества кластеров')

    # Сохраняем график
    image_path = 'silhouette_plot.png'
    plt.savefig(image_path)
    plt.close('all')  # Закрывает все открытые графики

    # Путь к существующему Excel-файлу "Отчёт"
    file_name = os.path.join(output_folder, 'Отчёт.xlsx')

    # Открываем существующий файл
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
    else:
        raise FileNotFoundError(f"Файл '{file_name}' не найден!")

    # Название нового листа для данных силуэта
    sheet_name = 'Силуэт'

    # Проверяем, существует ли уже лист "Силуэт", и создаём его, если нет
    if sheet_name not in workbook.sheetnames:
        worksheet_silhouette = workbook.create_sheet(sheet_name)
    else:
        worksheet_silhouette = workbook[sheet_name]

    # Записываем строку с оптимальным числом кластеров
    worksheet_silhouette.cell(row=1, column=1, value=f'Оптимальное количество кластеров: {best_n_clusters}')

    # Записываем заголовки для данных
    worksheet_silhouette.cell(row=3, column=1, value='Количество кластеров')
    worksheet_silhouette.cell(row=3, column=2, value='Silhouette Score')

    # Записываем данные анализа с помощью метода append
    for n_clusters, score in zip(x, m):
        worksheet_silhouette.append([n_clusters, score])

    # Вставляем график в Excel
    image_path = 'silhouette_plot.png'
    img = Image(image_path)
    worksheet_silhouette.add_image(img, 'C1')

    # Автовыравнивание ширины столбцов A и B
    max_len_col1 = max([len(str(n)) for n in x] + [len('Количество кластеров')])
    max_len_col2 = max([len(f'{s:.3f}') for s in m] + [len('Silhouette Score')])

    worksheet_silhouette.column_dimensions['A'].width = max_len_col1
    worksheet_silhouette.column_dimensions['B'].width = max_len_col2

    # Сохраняем изменения в файл Excel
    workbook.save(file_name)

    # Удаляем временные файлы изображений
    if os.path.exists(image_path):
        os.remove(image_path)

    print(f"График силуэта и результаты анализа успешно добавлены в '{file_name}' на лист '{sheet_name}'.")

    # Применение KMeans с оптимальным количеством кластеров
    if "Class" in scaled_df.columns:
        scaled_df.drop(columns=["Class"], inplace=True)

    km = KMeans(n_clusters=best_n_clusters, random_state=1000)

    scaled_df["Class"] = km.fit_predict(scaled_df)
    scaled_df["Class"] = scaled_df["Class"] + 1

    grouped_count = scaled_df.groupby("Class").count()["z-stat first"]
    print(grouped_count)

    temp["Class"] = scaled_df["Class"]

    mean_temp = temp.groupby("Class").mean()
    # Построение графика
    scaled_df.groupby("Class").mean().T.plot(grid=True, figsize=(15, 10),  # Увеличиваем высоту графика
                                             rot=90,
                                             xticks=range(len(scaled_df.columns) - 1),
                                             style='o-', linewidth=4, markersize=12)

    plt.legend(fontsize=30)
    plt.tick_params(axis='x', labelsize=14)
    plt.tick_params(axis='y', labelsize=10)  # Уменьшаем шрифт оси y
    plt.tight_layout()  # Подбираем отступы

    # Путь к Excel-файлу "Отчёт.xlsx"
    excel_file = os.path.join(output_folder, 'Отчёт.xlsx')

    # Сохраняем график
    image_path_z_stats = os.path.join(output_folder, 'z_stats.png')
    plt.savefig(image_path_z_stats)
    plt.close('all')  # Закрывает все открытые графики

    # Проверьте, что файл с графиком существует
    if not os.path.exists(image_path_z_stats):
        raise FileNotFoundError(f"Файл '{image_path_z_stats}' не был создан.")

    # Открываем существующий Excel-файл или создаём новый
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        workbook = Workbook()

    # Название нового листа
    sheet_name = 'Z-Статистики'

    # Проверяем, существует ли уже лист с названием "Статистики"
    if sheet_name not in workbook.sheetnames:
        worksheet_z_stats = workbook.create_sheet(sheet_name)
    else:
        worksheet_z_stats = workbook[sheet_name]

    # Вставляем таблицу grouped_count в Excel
    worksheet_z_stats.cell(row=1, column=1, value='Класс')
    worksheet_z_stats.cell(row=1, column=2, value='Число объектов')

    # Записываем данные из grouped_count
    row_num = 2
    for class_label, count in grouped_count.items():
        worksheet_z_stats.cell(row=row_num, column=1, value=class_label)
        worksheet_z_stats.cell(row=row_num, column=2, value=count)
        row_num += 1

    # Вставляем таблицу mean_temp (начиная с того же row_num)
    worksheet_z_stats.cell(row=row_num, column=1, value='Средние значения')
    row_num += 1

    # Записываем заголовки столбцов для mean_temp
    for col_num, col_name in enumerate(mean_temp.columns):
        worksheet_z_stats.cell(row=row_num, column=col_num + 2, value=col_name)  # Заголовки столбцов
    row_num += 1

    # Записываем значения для mean_temp
    for class_label, row in mean_temp.iterrows():
        worksheet_z_stats.cell(row=row_num, column=1, value=class_label)  # Записываем номер класса
        for col_num, value in enumerate(row):
            worksheet_z_stats.cell(row=row_num, column=col_num + 2, value=value)  # Записываем значения
        row_num += 1

    # Вставляем график в Excel
    img = Image(image_path_z_stats)
    img.width = img.width // 2.5  # Масштабирование изображения
    img.height = img.height // 2.5
    worksheet_z_stats.add_image(img, 'A8')

    # Устанавливаем автоподбор ширины для столбцов
    worksheet_z_stats.column_dimensions['A'].width = 15  # Ширина для столбца 'Класс'
    worksheet_z_stats.column_dimensions['B'].width = 20  # Ширина для столбца 'Число объектов'

    # Устанавливаем ширину для столбцов mean_temp
    for col_num in range(len(mean_temp.columns)):
        worksheet_z_stats.column_dimensions[chr(67 + col_num)].width = 20  # Столбцы с заголовками начинаются с C

    # Сохраняем изменения в файл Excel
    workbook.save(excel_file)

    # Удаляем временные файлы изображений
    if os.path.exists(image_path_z_stats):
        os.remove(image_path_z_stats)

    print(f"Лист 'Z-Статистики' добавлен к файлу '{excel_file}'. Файл готов для скачивания.")

    temp.groupby("Class").mean()

    # построение главных компонент

    from sklearn.decomposition import PCA
    update_log("Построение главных компонент...")
    pca = PCA(n_components=2)
    principalComponents = pca.fit_transform(scaled_df.iloc[:, :-1])
    principalDf = pd.DataFrame(data=principalComponents)

    principalDf = principalDf.rename(columns={0: "PC1", 1: "PC2"})
    principalDf["Class"] = scaled_df["Class"]

    principalDf.head()

    import pandas as pd
    import os
    import matplotlib.pyplot as plt
    from tqdm import tqdm
    from sklearn.cluster import KMeans
    import sklearn.metrics as metrics

    # Визуализация главных компонент
    for i in range(1, best_n_clusters + 1):
        data = principalDf[principalDf["Class"] == i]
        plt.plot(data.PC1, data.PC2, 'o', label=f'Класс {i}')
    plt.legend()
    plt.title("Главные компоненты")
    plt.xlabel("Главная компонента 1")
    plt.ylabel("Главная компонента 2")

    # Путь к Excel-файлу "Отчёт.xlsx"
    excel_file = os.path.join(output_folder, 'Отчёт.xlsx')

    # Сохраняем график
    image_path_pca = os.path.join(output_folder, 'pca_plot.png')
    plt.savefig(image_path_pca)
    plt.close('all')  # Закрывает все открытые графики

    # Проверьте, что файл с графиком существует
    if not os.path.exists(image_path_pca):
        raise FileNotFoundError(f"Файл '{image_path_pca}' не был создан.")

    # Открываем существующий Excel-файл или создаём новый, если файла нет
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        workbook = Workbook()

    # Название нового листа
    sheet_name = 'Метод ГК'

    # Проверяем, существует ли уже лист с названием "Метод ГК"
    if sheet_name not in workbook.sheetnames:
        worksheet_pca = workbook.create_sheet(sheet_name)
    else:
        worksheet_pca = workbook[sheet_name]

    # Вставляем график PCA в Excel
    img = Image(image_path_pca)
    worksheet_pca.add_image(img, 'A1')

    # Сохраняем изменения в файл Excel
    workbook.save(excel_file)

    # Удаляем временные файлы изображений
    if os.path.exists(image_path_pca):
        os.remove(image_path_pca)

    print(f"Лист 'Метод ГК' добавлен к файлу '{excel_file}'. Файл готов для скачивания.")

    # Анализ выбросов

    from sklearn.ensemble import IsolationForest
    update_log("Поиск аномального класса по Isolation Forest...")
    anomaly_labels = IsolationForest().fit_predict(scaled_df.drop(["Class"], axis=1))

    scaled_df["IsoLabels"] = anomaly_labels

    # Список для хранения отношений z-stat first
    ratios = {}

    # Получаем уникальные классы
    classes = scaled_df["Class"].unique()

    # Проходим по каждому классу и вычисляем отношение
    for cls in classes:
        # Получаем count z-stat first для текущего класса
        z_stat_pos = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["IsoLabels"] == 1)][
            "z-stat first"].count()
        z_stat_neg = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["IsoLabels"] == -1)][
            "z-stat first"].count()

        # Проверяем, чтобы z_stat_neg не было нулем
        if z_stat_neg != 0:
            ratio = z_stat_pos / z_stat_neg
            ratios[cls] = ratio
        else:
            ratio = z_stat_pos
            ratios[cls] = ratio

    # Находим класс с минимальным отношением
    if ratios:
        anomaly_class_iso = min(ratios, key=ratios.get)  # Класс с минимальным отношением
    else:
        anomaly_class_iso = None  # Если нет классов

    print(f'Аномальный класс по Isolation Forest: {anomaly_class_iso}')

    from sklearn.covariance import EllipticEnvelope
    update_log("Поиск аномального класса по Elliptic Envelope...")
    anomaly_labels_el = EllipticEnvelope().fit_predict(scaled_df.drop(["Class", "IsoLabels"], axis=1))

    scaled_df["ElLabels"] = anomaly_labels_el
    scaled_df.groupby(["ElLabels", "IsoLabels", "Class"]).count()

    ratios_el = {}

    # Проходим по каждому классу и вычисляем отношение для Elliptic Envelope
    for cls in classes:
        # Получаем count z-stat first для текущего класса
        z_stat_pos = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["ElLabels"] == 1)][
            "z-stat first"].count()
        z_stat_neg = scaled_df[(scaled_df["Class"] == cls) & (scaled_df["ElLabels"] == -1)][
            "z-stat first"].count()

        # Проверяем, чтобы z_stat_neg не было нулем
        if z_stat_neg != 0:
            ratio = z_stat_pos / z_stat_neg
            ratios_el[cls] = ratio
        else:
            ratio = z_stat_pos
            ratios_el[cls] = ratio

    # Находим аномальный класс для Elliptic Envelope
    if ratios_el:
        anomaly_class_el = min(ratios_el, key=ratios_el.get)  # Класс с минимальным отношением
    else:
        anomaly_class_el = None  # Если нет классов

    print(f'Аномальный класс по Elliptic Envelope: {anomaly_class_el}')

    import pandas as pd

    # Группируем данные и заполняем нулями
    grouped_iso = scaled_df.groupby(["IsoLabels", "Class"]).count().reindex(
        pd.MultiIndex.from_product([[-1, 1], classes], names=["IsoLabels", "Class"]), fill_value=0
    )

    grouped_el = scaled_df.groupby(["ElLabels", "Class"]).count().reindex(
        pd.MultiIndex.from_product([[-1, 1], classes], names=["ElLabels", "Class"]), fill_value=0
    )

    # Путь к Excel-файлу "Отчёт.xlsx"
    excel_file = os.path.join(output_folder, 'Отчёт.xlsx')

    # Открываем существующий Excel-файл или создаём новый, если файла нет
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
    else:
        workbook = Workbook()

    from openpyxl.utils import get_column_letter
    # Название нового листа
    sheet_name = 'Аномальный класс (этап 1)'

    # Проверяем, существует ли уже лист с названием "Аномальный класс (этап 1)"
    if sheet_name not in workbook.sheetnames:
        worksheet_iso = workbook.create_sheet(sheet_name)
    else:
        worksheet_iso = workbook[sheet_name]

    # Записываем заголовки для Isolation Forest
    worksheet_iso.cell(row=1, column=1, value='IsoLabels')
    worksheet_iso.cell(row=1, column=2, value='Класс')
    worksheet_iso.cell(row=1, column=3, value='Число объектов')

    # Записываем данные для IsoLabels
    for row_num, (index, row) in enumerate(grouped_iso.iterrows(), start=2):
        worksheet_iso.cell(row=row_num, column=1, value=index[0])  # IsoLabels
        worksheet_iso.cell(row=row_num, column=2, value=index[1])  # Class
        worksheet_iso.cell(row=row_num, column=3, value=row['z-stat first'])  # Count

    # Добавляем информацию об аномальном классе по Isolation Forest
    anomaly_row = len(grouped_iso) + 3
    worksheet_iso.cell(row=anomaly_row, column=1, value='Аномальный класс по Isolation Forest:')
    worksheet_iso.cell(row=anomaly_row, column=2, value=anomaly_class_iso)

    # Добавляем разделитель между результатами
    divider_row = anomaly_row + 2
    worksheet_iso.cell(row=divider_row, column=1, value='-----')  # Разделитель
    worksheet_iso.cell(row=divider_row + 1, column=1, value='Elliptic Envelope')  # Заголовок для Elliptic Envelope

    # Записываем заголовки для Elliptic Envelope
    elliptic_row_start = divider_row + 2
    worksheet_iso.cell(row=elliptic_row_start, column=1, value='ElLabels')
    worksheet_iso.cell(row=elliptic_row_start, column=2, value='Класс')
    worksheet_iso.cell(row=elliptic_row_start, column=3, value='Число объектов')

    # Записываем данные для ElLabels
    for row_num, (index, row) in enumerate(grouped_el.iterrows(), start=elliptic_row_start + 1):
        worksheet_iso.cell(row=row_num, column=1, value=index[0])  # ElLabels
        worksheet_iso.cell(row=row_num, column=2, value=index[1])  # Class
        worksheet_iso.cell(row=row_num, column=3, value=row['z-stat first'])  # Count

    # Добавляем информацию об аномальном классе по Elliptic Envelope
    worksheet_iso.cell(row=row_num + 1, column=1, value='Аномальный класс по Elliptic Envelope:')
    worksheet_iso.cell(row=row_num + 1, column=2, value=anomaly_class_el)

    # Автовыравнивание ширины столбцов
    for col in range(1, 4):
        column_letter = get_column_letter(col)
        worksheet_iso.column_dimensions[column_letter].width = 20

    # Сохраняем изменения в файл Excel
    workbook.save(excel_file)

    print(f"Лист 'Аномальный класс (этап 1)' добавлен к файлу '{excel_file}'.")

    scaled_df_temp = scaled_df.copy()
    scaled_df = scaled_df.drop(["ElLabels", "IsoLabels"], axis=1)

    from openpyxl.drawing.image import Image
    from openpyxl import load_workbook
    # Загружаем существующий файл Excel
    workbook = load_workbook(excel_file)

    # Название нового листа для графиков
    sheet_name = 'Аномальный класс (графики)'

    # Проверяем, существует ли уже лист с таким названием
    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        worksheet = workbook[sheet_name]

    # Создание графиков для каждого класса и сохранение в файлы PNG
    classes = scaled_df_temp["Class"].unique()
    image_files = []

    # Проходим по каждому классу и строим графики
    for cls in classes:
        plt.figure(figsize=(15, 8))

        # Основной график для среднего значения по классу
        scaled_df_temp[scaled_df_temp["Class"] == cls].iloc[:, :-3].mean().T.plot(
            style='o-', linewidth=4, markersize=12, label=f'Среднее по классу {cls}'
        )

        # График для IsoLabels=-1 и ElLabels=-1 для данного класса
        scaled_df_temp[
            (scaled_df_temp["Class"] == cls) &
            (scaled_df_temp["IsoLabels"] == -1) &
            (scaled_df_temp["ElLabels"] == -1)
            ].iloc[:, :-3].mean().T.plot(
            style='o--', linewidth=4, markersize=12, label=f'Аномалии по обоим методам (-1, -1)'
        )

        # График для IsoLabels=1 и ElLabels=1 для данного класса
        scaled_df_temp[
            (scaled_df_temp["Class"] == cls) &
            (scaled_df_temp["IsoLabels"] == 1) &
            (scaled_df_temp["ElLabels"] == 1)
            ].iloc[:, :-3].mean().T.plot(
            style='o--', linewidth=4, markersize=12, label=f'Нет аномалий по обоим методам (1, 1)'
        )

        # Настройки графика
        plt.legend(fontsize=14)
        plt.title(f'График для класса {cls}')
        plt.xlabel('Показатели')
        plt.ylabel('Значение')
        plt.grid(True, linestyle='--', linewidth=0.5)

        # Сохранение графика в PNG
        image_file = f'class_{cls}_plot.png'
        plt.savefig(image_file)
        image_files.append(image_file)
        plt.close('all')  # Закрывает все открытые графики

    # Координаты для вставки графиков
    start_row = 1  # Стартовая строка для первого графика

    # Вставка всех графиков на лист с уменьшением масштаба
    for cls, image_file in zip(classes, image_files):
        # Вставляем изображение графика в ячейку листа Excel
        img = Image(image_file)
        img.width, img.height = img.width * 0.5, img.height * 0.5  # Уменьшаем масштаб графиков
        worksheet.add_image(img, f'A{start_row}')  # Вставляем график в ячейку

        # Обновляем стартовую строку для следующего графика
        start_row += 20  # Подбираем это значение в зависимости от высоты графиков

    # Сохраняем изменения в файл Excel
    workbook.save(excel_file)

    # Удаляем временные файлы с графиками после вставки в Excel
    for image_file in image_files:
        os.remove(image_file)

    print(f'Графики успешно сохранены в файл {excel_file}')

    scaled_df = scaled_df_temp[(scaled_df_temp["ElLabels"] == -1) & (scaled_df_temp["IsoLabels"] == -1) & (
            scaled_df_temp["Class"] == anomaly_class_el)].iloc[:, :-2]

    scaled_df.head()

    # Получаем индексы аномальных объектов из scaled_df
    anomaly_indices_el = scaled_df[scaled_df["Class"] == anomaly_class_el].index

    # Фильтрация исходного DataFrame df по этим индексам
    anomaly_original_rows = df.loc[anomaly_indices_el]

    # Вывод
    anomaly_original_rows.head(10)

    # Название нового листа, на который нужно сохранить данные
    sheet_name = 'Подозрительные операции'
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Записываем данные на новый лист
        anomaly_original_rows.to_excel(writer, sheet_name=sheet_name, index=False)

    cluster_1_index = list(scaled_df[scaled_df["Class"] == anomaly_class_el].index)

    # df_cur[(df_cur.index.ster_1_isin(cluindex)) & (df_cur["ВидСубконтоДт1"] == "Контрагенты")].groupby("СубконтоДт1").count().sort_values(by="Сумма", ascending = False).head(10)

    df_class_1 = temp[(temp.index.isin(cluster_1_index))]

    temp_class_1 = df_class_1.loc[:,
                   ["z-stat first", "z-stat second", "z-stat first_two", "sum_frequency", "z_stat_second_diff",
                    "Частота суммы", "Частота счета Дт", "Частота счета Кт", "Частота проводки",
                    "Частота автора операции",
                    "Ручная проводка", "Выходные или рабочие", "Сторно", "Количество дублей"]]

    """**3.2 Кластеризация, Этап 2**"""

    # стандартизация данных
    # проведем кластеризацию для объектов аномального класса по оставшимся признакам
    update_log("Построение силуэта аномального класса...")
    from sklearn.preprocessing import StandardScaler

    scaled_ = StandardScaler().fit_transform(temp_class_1.iloc[:, 6:].values)
    scaled_class_1 = pd.DataFrame(scaled_, index=temp_class_1.iloc[:, 6:].index,
                                  columns=temp_class_1.iloc[:, 6:].columns)
    scaled_class_1.head()

    # Силуэт
    plt.close('all')  # Закрывает все открытые графики

    plt.rcParams["figure.figsize"] = (8, 4)

    x = [i for i in range(2, 6)]
    m = []

    sample_size = int(scaled_class_1.shape[0])
    print(f'sample size = {sample_size}')
    for i in tqdm(x):
        print(i)
        labels = KMeans(n_clusters=i, random_state=1000).fit(scaled_class_1).labels_
        print(labels)
        m.append(metrics.silhouette_score(scaled_class_1, labels, sample_size=sample_size))
        print(f'n = {i}, silhouette = {m[-1]}')
    # Найдём количество кластеров с максимальным значением silhouette_score
    best_n_clusters = x[m.index(max(m))]

    # Строим график зависимости silhouette_score от количества кластеров
    plt.plot(x, m, 'r-')
    plt.xticks(ticks=x, labels=[int(i) for i in x])
    plt.xlabel('Количество кластеров')
    plt.ylabel('Значение метрики')
    plt.title('Зависимость значения метрики от количества кластеров')

    # Сохраняем график
    image_path = 'silhouette_plot_2.png'
    plt.savefig(image_path)
    plt.close('all')  # Закрывает все открытые графики

    # Путь к существующему Excel-файлу "Отчёт"
    file_name = os.path.join(output_folder, 'Отчёт.xlsx')

    # Открываем существующий файл
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
    else:
        raise FileNotFoundError(f"Файл '{file_name}' не найден!")

    # Название нового листа для данных силуэта
    sheet_name = 'Силуэт (этап 2)'

    # Проверяем, существует ли уже лист "Силуэт", и создаём его, если нет
    if sheet_name not in workbook.sheetnames:
        worksheet_silhouette = workbook.create_sheet(sheet_name)
    else:
        worksheet_silhouette = workbook[sheet_name]

    # Записываем строку с оптимальным числом кластеров
    worksheet_silhouette.cell(row=1, column=1, value=f'Оптимальное количество кластеров: {best_n_clusters}')

    # Записываем заголовки для данных
    worksheet_silhouette.cell(row=3, column=1, value='Количество кластеров')
    worksheet_silhouette.cell(row=3, column=2, value='Silhouette Score')

    # Записываем данные анализа с помощью метода append
    for n_clusters, score in zip(x, m):
        worksheet_silhouette.append([n_clusters, score])

    img = Image(image_path)
    worksheet_silhouette.add_image(img, 'C1')

    # Автовыравнивание ширины столбцов A и B
    max_len_col1 = max([len(str(n)) for n in x] + [len('Количество кластеров')])
    max_len_col2 = max([len(f'{s:.3f}') for s in m] + [len('Silhouette Score')])

    worksheet_silhouette.column_dimensions['A'].width = max_len_col1
    worksheet_silhouette.column_dimensions['B'].width = max_len_col2

    # Сохраняем изменения в файл Excel
    workbook.save(file_name)

    # Удаляем временные файлы изображений
    if os.path.exists(image_path):
        os.remove(image_path)

    print(f"График силуэта и результаты анализа успешно добавлены в '{file_name}' на лист '{sheet_name}'.")
    plt.close('all')  # Закрывает все открытые графики

    n_clusters = best_n_clusters
    if "Class" in scaled_class_1.columns:
        scaled_class_1.drop(columns=["Class", ], inplace=True)
    km = KMeans(n_clusters=n_clusters, random_state=1000)
    scaled_class_1["Class"] = km.fit_predict(scaled_class_1)
    scaled_class_1["Class"] = scaled_class_1["Class"] + 1

    import pandas as pd
    import matplotlib.pyplot as plt
    from sklearn.cluster import KMeans

    # Удаляем столбец "Class", если он уже есть
    if "Class" in scaled_class_1.columns:
        scaled_class_1.drop(columns=["Class"], inplace=True)

    # Применение KMeans
    km = KMeans(n_clusters=n_clusters, random_state=1000)
    scaled_class_1["Class"] = km.fit_predict(scaled_class_1)
    scaled_class_1["Class"] = scaled_class_1["Class"] + 1  # Нумерация классов с 1

    # Вычисляем количество объектов в каждом классе
    class_counts = scaled_class_1.groupby("Class").count()["Сторно"]
    class_counts = class_counts.reset_index()  # Преобразуем индекс в столбец для удобного вывода

    # Переименовываем столбцы
    class_counts.columns = ['Класс', 'Количество объектов']

    # Рисуем график средних значений по каждому классу
    scaled_class_1.groupby("Class").mean().T.plot(grid=True, figsize=(15, 10), rot=90,
                                                  xticks=range(len(scaled_class_1.columns) - 1), style='o-',
                                                  linewidth=4,
                                                  markersize=12)

    plt.legend(fontsize=30)
    plt.tick_params(axis='x', labelsize=18)
    plt.tick_params(axis='y', labelsize=18)

    # Название нового листа для кластеризации
    sheet_name = 'Кластеризация (этап 2)'

    # Сохраняем график во временный файл
    image_path = 'cluster_plot.png'
    plt.savefig(image_path, bbox_inches='tight')
    plt.close('all')  # Закрываем все открытые графики

    # Открываем существующий файл Excel
    workbook = load_workbook(excel_file)

    # Проверяем, существует ли уже лист с таким названием
    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        worksheet = workbook[sheet_name]

    # Добавляем заголовки
    worksheet.cell(row=1, column=1, value="Класс")
    worksheet.cell(row=1, column=2, value="Число объектов")

    # Записываем данные по количеству объектов в каждом классе в Excel
    for idx, row in class_counts.iterrows():
        worksheet.append(row.tolist())  # Записываем каждую строку данных

    # Вставляем график ниже данных
    start_row = worksheet.max_row + 5  # Позиционируем график через 5 строк после последней записи
    img = Image(image_path)
    img.width, img.height = img.width * 0.7, img.height * 0.7  # Уменьшаем масштаб графика
    worksheet.add_image(img, f'A{start_row}')

    # Сохраняем изменения в файл Excel
    workbook.save(excel_file)

    # Удаляем временный файл с графиком
    if os.path.exists(image_path):
        os.remove(image_path)

    print(f'График и данные кластеров успешно добавлены в файл {excel_file}.')

    # Анализ выбросов
    update_log("Поиск аномального класса по Isolation Forest...")
    from sklearn.ensemble import IsolationForest

    anomaly_labels = IsolationForest().fit_predict(scaled_class_1.drop(["Class"], axis=1))

    import pandas as pd

    scaled_class_1["IsoLabels"] = anomaly_labels

    # Список для хранения отношений Сторно
    ratios = {}

    # Определяем порядок классов
    class_order = sorted(scaled_class_1["Class"].unique())  # [1, 2, 3, 4, 5]

    # Проходим по каждому классу и вычисляем отношение
    for cls in class_order:
        z_stat_pos = scaled_class_1[(scaled_class_1["Class"] == cls) & (scaled_class_1["IsoLabels"] == 1)][
            "Сторно"].count()
        z_stat_neg = scaled_class_1[(scaled_class_1["Class"] == cls) & (scaled_class_1["IsoLabels"] == -1)][
            "Сторно"].count()

        if z_stat_neg != 0:
            ratio = z_stat_pos / z_stat_neg
            ratios[cls] = ratio
        else:
            ratio = z_stat_pos
            ratios[cls] = ratio

    # Находим класс с минимальным отношением
    if ratios:
        anomaly_class_iso = min(ratios, key=ratios.get)
    else:
        anomaly_class_iso = None

    print(f'Аномальный класс по Isolation Forest: {anomaly_class_iso}')

    from sklearn.covariance import EllipticEnvelope
    update_log("Поиск аномального класса по Elliptic Envelope...")
    anomaly_labels_el = EllipticEnvelope().fit_predict(scaled_class_1.drop(["Class", "IsoLabels"], axis=1))

    scaled_class_1["ElLabels"] = anomaly_labels_el

    ratios_el = {}

    for cls in class_order:
        z_stat_pos = scaled_class_1[(scaled_class_1["Class"] == cls) & (scaled_class_1["ElLabels"] == 1)][
            "Сторно"].count()
        z_stat_neg = scaled_class_1[(scaled_class_1["Class"] == cls) & (scaled_class_1["ElLabels"] == -1)][
            "Сторно"].count()

        if z_stat_neg != 0:
            ratio = z_stat_pos / z_stat_neg
            ratios_el[cls] = ratio
        else:
            ratio = z_stat_pos
            ratios_el[cls] = ratio

    if ratios_el:
        anomaly_class_el = min(ratios_el, key=ratios_el.get)
    else:
        anomaly_class_el = None

    print(f'Аномальный класс по Elliptic Envelope: {anomaly_class_el}')

    # Группируем данные и заполняем нулями, устанавливая порядок классов
    grouped_iso = scaled_class_1.groupby(["IsoLabels", "Class"]).count().reindex(
        pd.MultiIndex.from_product([[-1, 1], class_order], names=["IsoLabels", "Class"]), fill_value=0
    )

    grouped_el = scaled_class_1.groupby(["ElLabels", "Class"]).count().reindex(
        pd.MultiIndex.from_product([[-1, 1], class_order], names=["ElLabels", "Class"]), fill_value=0
    )

    # Название нового листа
    sheet_name = 'Аномальный класс (этап 2)'

    # Открываем существующий Excel файл
    workbook = load_workbook(excel_file)

    # Проверяем, существует ли уже лист с названием "Аномальный класс (этап 2)"
    if sheet_name not in workbook.sheetnames:
        worksheet_iso = workbook.create_sheet(sheet_name)
    else:
        worksheet_iso = workbook[sheet_name]

    # Записываем заголовки для Isolation Forest
    worksheet_iso.cell(row=1, column=1, value='IsoLabels')
    worksheet_iso.cell(row=1, column=2, value='Класс')
    worksheet_iso.cell(row=1, column=3, value='Число объектов')

    # Записываем данные для IsoLabels
    for row_num, (index, row) in enumerate(grouped_iso.iterrows(), start=2):
        worksheet_iso.cell(row=row_num, column=1, value=index[0])  # IsoLabels
        worksheet_iso.cell(row=row_num, column=2, value=index[1])  # Class
        worksheet_iso.cell(row=row_num, column=3, value=row['Сторно'])  # Count

    # Добавляем информацию об аномальном классе по Isolation Forest
    anomaly_row = len(grouped_iso) + 3
    worksheet_iso.cell(row=anomaly_row, column=1, value='Аномальный класс по Isolation Forest:')
    worksheet_iso.cell(row=anomaly_row, column=2, value=anomaly_class_iso)

    # Добавляем разделитель между результатами
    divider_row = anomaly_row + 2
    worksheet_iso.cell(row=divider_row, column=1, value='-----')  # Разделитель
    worksheet_iso.cell(row=divider_row + 1, column=1, value='Elliptic Envelope')  # Заголовок для Elliptic Envelope

    # Записываем заголовки для Elliptic Envelope
    elliptic_row_start = divider_row + 2
    worksheet_iso.cell(row=elliptic_row_start, column=1, value='ElLabels')
    worksheet_iso.cell(row=elliptic_row_start, column=2, value='Класс')
    worksheet_iso.cell(row=elliptic_row_start, column=3, value='Число объектов')

    # Записываем данные для ElLabels
    for row_num, (index, row) in enumerate(grouped_el.iterrows(), start=elliptic_row_start + 1):
        worksheet_iso.cell(row=row_num, column=1, value=index[0])  # ElLabels
        worksheet_iso.cell(row=row_num, column=2, value=index[1])  # Class
        worksheet_iso.cell(row=row_num, column=3, value=row['Сторно'])  # Count

    # Добавляем информацию об аномальном классе по Elliptic Envelope
    worksheet_iso.cell(row=row_num + 1, column=1, value='Аномальный класс по Elliptic Envelope:')
    worksheet_iso.cell(row=row_num + 1, column=2, value=anomaly_class_el)

    # Автовыравнивание ширины столбцов
    from openpyxl.utils import get_column_letter

    for col in range(1, 4):
        column_letter = get_column_letter(col)
        worksheet_iso.column_dimensions[column_letter].width = 20

    # Сохраняем изменения в файл Excel
    workbook.save(excel_file)

    print(f"Лист 'Аномальный класс (этап 2)' добавлен к файлу '{excel_file}'.")

    # Вывод подозрительных операций в новый лист Excel
    update_log("Вывод подозрительных операций...")
    # Получаем индексы аномальных объектов из scaled_class_1
    anomaly_indices_el = scaled_class_1[scaled_class_1["Class"] == anomaly_class_el].index

    # Фильтрация исходного DataFrame df по этим индексам
    anomaly_original_rows = df.loc[anomaly_indices_el]

    # Сохраняем аномальные строки в существующий файл на новый лист
    sheet_name = 'Подозрительные операции (итог)'
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Записываем данные на новый лист
        anomaly_original_rows.to_excel(writer, sheet_name=sheet_name, index=False)

    # Сообщение о завершении работы
    messagebox.showinfo("Результат", f"Тест завершен. Проверьте вывод в папку {output_folder} для деталей.")

    # После завершения анализа скрываем все кнопки, кроме "Выход" и "Провести новый анализ"
    start_button.pack_forget()
    load_osv_button.pack_forget()
    start_analysis_button.pack_forget()
    update_log("Анализ завершён.")
    log_label.pack_forget()  # Можно скрыть метку, если она больше не нужна

    # Показать кнопку "Провести новый анализ"
    restart_button.pack(side=tk.LEFT, padx=10, pady=10)


# Функция для выхода из программы
def exit_program():
    window.quit()


# Функция для перезапуска окна приложения
def restart_program():
    window.destroy()
    create_main_window()


def create_main_window():
    global window, log_label, load_osv_button, start_analysis_button, start_button, restart_button, loading_label
    button_font = ('Helvetica', 12, "bold")
    # Создаем основное окно
    window = tk.Tk()
    window.title("Анализ данных")
    window.geometry("400x300")
    window.configure(bg="#f0f0f0")  # Изменяем цвет фона окна

    # Кнопка для загрузки первого файла
    start_button = tk.Button(window, text="Загрузить первый файл (ЖО)", command=load_first_file, bg="#FFFFFF",
                             fg="blue", font=button_font)
    start_button.pack(pady=10, padx=10, fill='x')  # Увеличиваем отступы и заполняем по горизонтали

    # Кнопка для загрузки второго файла
    load_osv_button = tk.Button(window, text="Загрузить второй файл (ОСВ)", command=load_second_file, bg="#FFFFFF",
                                fg="blue", font=button_font)
    load_osv_button.pack(pady=10, padx=10, fill='x')
    load_osv_button.config(state=tk.DISABLED)

    # Кнопка для начала анализа
    start_analysis_button = tk.Button(window, text="Начать анализ", command=start_analysis, bg="#FFFFFF", fg="blue",
                                      font=button_font)
    start_analysis_button.pack(pady=10, padx=10, fill='x')
    start_analysis_button.pack_forget()

    # Лог для отображения текущих действий
    log_label = tk.Label(window, text="Лог действий будет здесь", anchor="w", bg="#f0f0f0", font=button_font)
    log_label.pack(side="bottom", fill="x", padx=10, pady=10)

    # Кнопка для проведения нового анализа (изначально скрыта)
    restart_button = tk.Button(window, text="Провести новый анализ", command=restart_program, bg="#FFFFFF", fg="blue",
                               font=button_font)
    restart_button.pack_forget()  # Изначально скрыта

    # Кнопка выхода
    exit_button = tk.Button(window, text="Выход", command=exit_program, bg="#f44336", fg="white", font=button_font)
    exit_button.pack(side="right", padx=10, pady=10)

    window.mainloop()


# Запуск приложения
create_main_window()
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import pandas as pd
from datetime import datetime

class ExcelWriter:

    def __init__(self, output_folder, company_name, filename=None):
        """
        Инициализация класса для работы с Excel.

        :param output_folder: Путь к папке для сохранения файлов
        :param filename: Название файла
        """
        self.output_folder = output_folder
        self.company_name = company_name
        current_date = datetime.now().strftime('%d-%m-%Y')

        # Если filename не передан, формируем его автоматически
        if not filename:
            self.filename = f"{self.company_name}_{current_date}.xlsx"
        else:
            self.filename = f"{filename}.xlsx"

        self.file_path = os.path.join(self.output_folder, self.filename)
        # Создание папки, если она не существует
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

        # Удаляем файл Excel, если он уже существует
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
            print(f"Файл {self.file_path} удален.")

        self.book = Workbook()  # Инициализация новой рабочей книги

    def remove_default_sheet(self):
        """
        Удаляет дефолтный лист 'Sheet', если он существует и является пустым.
        """
        if 'Sheet' in self.book.sheetnames and len(self.book.sheetnames) == 1:
            sheet = self.book['Sheet']
            if not sheet.max_row > 1:  # Проверка, пуст ли лист
                self.book.remove(sheet)

    def save_data_to_excel(self, df, sheet_name):
        """
        Добавляет данные на новый лист в Excel.

        :param df: DataFrame с результатом теста
        :param sheet_name: Название листа для записи данных
        """
        # Удаление дефолтного листа, если он есть
        self.remove_default_sheet()

        if sheet_name not in self.book.sheetnames:
            sheet = self.book.create_sheet(sheet_name)
        else:
            sheet = self.book[sheet_name]

        # Записываем данные DataFrame на лист, начиная с первой строки
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Явная очистка ячейки A1
        if sheet["A1"].value is not None:
            sheet["A1"].value = None

        self.save()

    def save_data_with_charts(self, df, sheet_name, chart_paths=None):
        """
        Добавляет данные и, если предоставлены, график(и) на новый лист в Excel.

        :param df: DataFrame с результатом теста
        :param sheet_name: Название листа для записи данных
        :param chart_paths: Путь к изображению графика или список путей к графикам (опционально)
        """
        # Убедимся, что chart_paths — это список. Если это не список и не None, преобразуем его в список.
        if chart_paths and not isinstance(chart_paths, list):
            chart_paths = [chart_paths]

        # Сохраняем данные на лист
        self.save_data_to_excel(df, sheet_name)

        # Если предоставлены пути к графикам, добавляем их
        if chart_paths:
            self.add_charts(chart_paths, sheet_name)

    def add_charts(self, chart_paths, sheet_name, position=None):
        """
        Добавляет график или несколько графиков на лист Excel.

        :param chart_paths: Список путей к графикам
        :param sheet_name: Лист, на который нужно вставить график
        :param position: Позиция графика на листе (по умолчанию смещено на 5 ячеек вправо от последнего столбца)
        """
        sheet = self.book[sheet_name]
        index = 5  # Начнем с 5-й строки для добавления изображений

        for chart_path in chart_paths:
            if os.path.exists(chart_path):
                img = Image(chart_path)

                # Уменьшение размера изображения до половины
                img.width //= 2.5
                img.height //= 2.5

                # Если позиция не указана, смещаем на 5 ячеек вправо от последнего столбца
                if not position:
                    max_column = sheet.max_column
                    image_column = max_column + 5  # Смещаем на 5 столбцов вправо
                    position = f'{chr(65 + image_column)}{index}'  # Начинаем с 5-й строки и двигаемся вниз

                sheet.add_image(img, position)
                index += 15  # Сместимся вниз на 15 строк для следующего графика
                self.save()
            else:
                print(f"Файл изображения '{chart_path}' не найден.")

    def delete_png_files(self):
        """Удаляет все .png файлы в директории, где находится исполняемый файл."""
        for file_name in os.listdir(os.getcwd()):
            if file_name.endswith('.png'):
                os.remove(file_name)
                print(f"Файл изображения {file_name} удален.")

    def save(self):
        """ Сохраняет книгу на диск. """
        self.book.save(self.file_path)
        print(f'Файл сохранен: {self.file_path}')
